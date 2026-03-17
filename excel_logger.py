# excel_logger.py
# ============================================================
# Smart Atlas — Excel Data Logger
# Appends form submissions to smart_atlas_data.xlsx
#
# Sheets & Columns:
#   Signup_Data    → Name, Email, Password, Date
#   Login_Data     → Email, Login_Time
#   Itinerary_Data → User_Email, Destination, Travel_Date, Budget, Notes
#   Feedback_Data  → User_Email, Rating, Comment, Date
#
# Usage (from any module):
#   import excel_logger as xl
#   xl.log_signup(name, email, password_hash)
#   xl.log_login(email)
#   xl.log_itinerary(email, destination, travel_date, budget, notes)
#   xl.log_feedback(email, rating, comment)
# ============================================================

import logging
import os
from datetime import datetime, date

logger = logging.getLogger(__name__)

# ── Excel file path (same folder as log_p.py) ─────────────────
_EXCEL_PATH = "smart_atlas_data.xlsx"

# ── Sheet names and their header rows ─────────────────────────
_SHEETS = {
    "Signup_Data":    ["Name", "Email", "Password", "Date"],
    "Login_Data":     ["Email", "Login_Time"],
    "Itinerary_Data": ["User_Email", "Destination", "Travel_Date", "Budget", "Notes"],
    "Feedback_Data":  ["User_Email", "Rating", "Comment", "Date"],
}

# ── Column widths for readability ──────────────────────────────
_COL_WIDTHS = {
    "Signup_Data":    [20, 30, 30, 20],
    "Login_Data":     [30, 22],
    "Itinerary_Data": [30, 22, 16, 12, 40],
    "Feedback_Data":  [30, 10, 50, 20],
}


def _get_workbook():
    """
    Load existing workbook or create a brand-new one.
    Ensures all 4 sheets exist with headers.
    Returns the openpyxl Workbook object.
    """
    try:
        from openpyxl import load_workbook, Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError(
            "openpyxl not installed. Run: pip install openpyxl"
        )

    if os.path.exists(_EXCEL_PATH):
        wb = load_workbook(_EXCEL_PATH)
    else:
        wb = Workbook()
        # Remove the default empty sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # Header style
    header_font    = Font(bold=True, color="FFFFFF", size=11)
    header_fill    = PatternFill(start_color="1B3B8B", end_color="1B3B8B", fill_type="solid")
    header_align   = Alignment(horizontal="center", vertical="center")
    thin_border    = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    # Ensure every sheet exists with correct headers
    for sheet_name, headers in _SHEETS.items():
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            ws.row_dimensions[1].height = 20

            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font    = header_font
                cell.fill    = header_fill
                cell.alignment = header_align
                cell.border  = thin_border

            # Set column widths
            widths = _COL_WIDTHS.get(sheet_name, [])
            for col_idx, width in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Keep the 4 sheets in the correct order
    desired_order = list(_SHEETS.keys())
    current_sheets = wb.sheetnames

    # Move sheets into correct order if they exist
    for idx, name in enumerate(desired_order):
        if name in current_sheets:
            wb.move_sheet(name, offset=wb.sheetnames.index(name) - idx)

    return wb


def _append_row(sheet_name: str, row_data: list) -> bool:
    """
    Thread-safe append of one row to the given sheet.
    Returns True on success, False on failure.
    """
    try:
        from openpyxl.styles import Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = _get_workbook()
        ws = wb[sheet_name]

        # Find the next empty row
        next_row = ws.max_row + 1

        # Zebra stripe: alternate row colors
        from openpyxl.styles import PatternFill
        row_fill = PatternFill(
            start_color="EEF2FF" if next_row % 2 == 0 else "FFFFFF",
            end_color="EEF2FF"   if next_row % 2 == 0 else "FFFFFF",
            fill_type="solid",
        )
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"),  bottom=Side(style="thin"),
        )
        center_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=next_row, column=col_idx, value=value)
            cell.fill      = row_fill
            cell.border    = thin_border
            cell.alignment = center_align

        wb.save(_EXCEL_PATH)
        logger.info("excel_logger: wrote to %s row %d", sheet_name, next_row)
        return True

    except Exception as e:
        logger.error("excel_logger: _append_row failed for %s: %s", sheet_name, e)
        return False


# ══════════════════════════════════════════════════════════════
# PUBLIC API — one function per form
# ══════════════════════════════════════════════════════════════

def log_signup(name: str, email: str, password_hash: str) -> bool:
    """
    Called after a new user is successfully created in MySQL.
    Writes to Sheet1 → Signup_Data.

    Args:
        name          : user's full name
        email         : user's email
        password_hash : bcrypt hash (never plain-text)
    """
    row = [
        name.strip()          if name  else "",
        email.strip().lower() if email else "",
        password_hash[:20] + "..." if password_hash else "",   # truncate for safety
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    ]
    return _append_row("Signup_Data", row)


def log_login(email: str) -> bool:
    """
    Called after a successful password check in login_page().
    Writes to Sheet2 → Login_Data.

    Args:
        email : the email that just logged in
    """
    row = [
        email.strip().lower() if email else "",
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    ]
    return _append_row("Login_Data", row)


def log_itinerary(
    email: str,
    destination: str,
    travel_date: str = "",
    budget: str = "",
    notes: str = "",
) -> bool:
    """
    Called after 'Generate Itinerary' succeeds.
    Writes to Sheet3 → Itinerary_Data.

    Args:
        email       : logged-in user's email
        destination : city/country chosen
        travel_date : start date string (optional)
        budget      : 'Budget' or 'Luxury' (optional)
        notes       : extra info e.g. duration + transport (optional)
    """
    row = [
        email.strip().lower() if email       else "",
        destination.strip()   if destination else "",
        str(travel_date)      if travel_date else datetime.now().strftime("%Y-%m-%d"),
        budget.strip()        if budget      else "",
        notes.strip()         if notes       else "",
    ]
    return _append_row("Itinerary_Data", row)


def log_feedback(
    email: str,
    rating: int,
    comment: str = "",
) -> bool:
    """
    Called after feedback form is submitted successfully.
    Writes to Sheet4 → Feedback_Data.

    Args:
        email   : user's email (from form input or session)
        rating  : overall_rating integer (1–5)
        comment : combined suggestions / expectations text
    """
    row = [
        email.strip().lower() if email   else "anonymous",
        int(rating)           if rating  else 0,
        comment.strip()       if comment else "",
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    ]
    return _append_row("Feedback_Data", row)


def get_all_data() -> dict:
    """
    Read all sheets back as lists of dicts.
    Useful for displaying data in the app.
    Returns: { sheet_name: [{ col: value, ... }, ...] }
    """
    result = {}
    try:
        from openpyxl import load_workbook
        if not os.path.exists(_EXCEL_PATH):
            return {}
        wb = load_workbook(_EXCEL_PATH)
        for sheet_name in _SHEETS:
            if sheet_name not in wb.sheetnames:
                result[sheet_name] = []
                continue
            ws     = wb[sheet_name]
            rows   = list(ws.iter_rows(values_only=True))
            if not rows:
                result[sheet_name] = []
                continue
            headers = rows[0]
            result[sheet_name] = [
                {headers[i]: row[i] for i in range(len(headers))}
                for row in rows[1:]
                if any(cell is not None for cell in row)
            ]
    except Exception as e:
        logger.error("excel_logger: get_all_data failed: %s", e)
    return result
